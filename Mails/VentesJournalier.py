# https://www.geeksforgeeks.org/send-mail-attachment-gmail-account-using-python/

# libraries to be imported
import pyodbc
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import datetime as dt, time
import os, glob
from openpyxl import Workbook, load_workbook

today = dt.date.today()
location = r"E:\SQLPython\Ventes\{0}-{1}{2}".format("Vente", today, ".xlsx")
nomFichier ="{0}-{1}{2}".format("Vente", today, ".xlsx")

# Create workbook
wb = Workbook()
# Add sheets to workbook
# Designate sheet name and position
sheet1 = wb.create_sheet('Abonnement',0)
sheet2 = wb.create_sheet('RO',1)
sheet3 = wb.create_sheet('RS',2)
sheet4 = wb.create_sheet('Migration',3)


def connexion():

  #connexion avec la bdd db_contact
  conn = pyodbc.connect(Driver = '{SQL Server Native Client 11.0}',
                        Server = 'SRV-DB',
                        Database = 'SCLDBV1',
                        user='sa',
                        password="Scl@2016"
                        )
  return conn


#-- ABONNEMENT
def abonnement(date):
    cnx = connexion()
    # prepare a cursor object using cursor() method
    cursor = cnx.cursor()
    cursor.execute(
        "select cc.CON_NO, om.ordersid as 'Facture n°', om.total_amount as 'montant total', om.PAY_TYPE "\
        "as 'type réglement 1=Espece | 2=Cheque',om.energy, sum(ofp.fixed_energy) as 'composante energetique',  sum(ofp.charge_of_table) as 'charge de tableau', sum(ofp.fer_fee) 'FER', "\
        "sum(ofp.vat_fee) 'TVA', sum(ofp.tco_fee) 'TCO',FORMAT(om.op_time,'yyyy-MM-dd') as 'date opération', om.operator "\
        "from ORDER_MASTER om inner join Order_fixed_package ofp on om.ORDERSID=ofp.ORDERSID inner join CON_CONTRACT cc on om.contractNo=cc.CON_NO "\
        "where  format(om.OP_TIME,'yyyy-MM-dd')= '"+date+"' "\
        "AND om.ORDER_TYPE <> 02 and om.ORDER_TYPE=26 and POSID=1549 "\
        "group by cc.CON_NO, om.ordersid, om.order_type, om.total_amount, om.operator, om.energy, om.account_pay_amount, om.actual_pay_amount,om.OP_TIME, om.PAY_TYPE " \
        "order by om.OPERATOR")

    # num_fields = len(cursor.description)
    field_names = [i[0] for i in cursor.description]
    # Mettre dans une liste
    results = cursor.fetchall()
    # Parcourir et print
    active = wb['Abonnement']
    active.append(field_names)
    for row in results:
        # make a row as List
        listrow = list(row)

        active.append(listrow)
    active = wb['Sheet']
    active['A1'] = "Abonnement"
    active['K1'] = "Migration"
    active['H1'] = "RS"
    active['E1'] = "RO"
    wb.save(location)
    print("Success!!!! Abonnement done")
    return results


def rechargeObligatoire(date):
    cnx = connexion()
    # prepare a cursor object using cursor() method
    cursor = cnx.cursor()
    cursor.execute("select cc.CON_NO, om.ordersid as 'Facture n°', om.total_amount as 'montant total', ISNULL(om.account_pay_amount,0) as 'payement du solde', "\
                  "ISNULL(om.actual_pay_amount,0) as 'payé', om.PAY_TYPE as 'type réglement 1=Espece | 2=Cheque', "\
                  "om.energy, sum(ofp.fixed_energy) as 'composante energetique',  "\
                  "sum(ofp.charge_of_table) as 'charge de tableau', sum(ofp.fer_fee) 'FER', "\
                  "sum(ofp.vat_fee) 'TVA', sum(ofp.tco_fee) 'TCO',FORMAT(om.op_time,'yyyy-MM-dd') as 'date opération', om.operator "\
                  "from ORDER_Master om "\
                  "inner join Order_fixed_package ofp on om.ORDERSID=ofp.ORDERSID "\
                  "inner join CON_CONTRACT cc on om.contractNo=cc.CON_NO  "\
                  "where   format(om.OP_TIME,'yyyy-MM-dd')= '"+date+"' AND om.ORDER_TYPE <> 02 and om.ORDER_TYPE=01 and POSID=1549 "\
                  "group by cc.CON_NO, om.ordersid, om.order_type, om.total_amount, om.operator, "\
                  "om.energy, om.account_pay_amount, om.actual_pay_amount,om.OP_TIME, om.PAY_TYPE order by om.OPERATOR")

    field_names = [i[0] for i in cursor.description]
    # Mettre dans une liste
    results = cursor.fetchall()
    # Parcourir et print
    active = wb['RO']
    active.append(field_names)
    for row in results:
        # make a row as List
        listrow = list(row)

        active.append(listrow)
    wb.save(location)
    print("Success!!!! RO done")
    return results

def rechargeSupplementaire(date):
    cnx = connexion()
    # prepare a cursor object using cursor() method
    cursor = cnx.cursor()
    cursor.execute("SELECT contractNo, TI as 'service', TOTAL_AMOUNT as 'montant total',om.account_pay_amount as 'payement du solde', "\
                  "om.actual_pay_amount as 'payé',om.PAY_TYPE as 'type réglement 1=Espece | 2=Cheque', "\
                  "ADDTIONAL_ENERGY,ADDTIONAL_ENERGY_AMOUNT,sum(item_ammount) as 'frais', "\
                  "FORMAT(om.op_time,'yyyy-MM-dd') as 'date opération',om.OPERATOR FROM ORDER_MASTER om "\
                  "inner join ORDER_CHARGE oc on om.ORDERSID=oc.ORDERSID WHERE VENDINGMETHOD=02 and  format(om.OP_TIME,'yyyy-MM-dd')= '"+date+"' and ITEM_AMMOUNT>0 and POSID=1549 "\
                  "group by contractNo, TI, TOTAL_AMOUNT,om.account_pay_amount, om.actual_pay_amount, ADDTIONAL_ENERGY,ADDTIONAL_ENERGY_AMOUNT,om.OP_TIME, om.OPERATOR,om.PAY_TYPE "\
                  "order by om.OP_TIME")

    # num_fields = len(cursor.description)
    field_names = [i[0] for i in cursor.description]
    # Mettre dans une liste
    results = cursor.fetchall()
    # Parcourir et print
    active = wb['RS']
    active.append(field_names)
    for row in results:
        # make a row as List
        listrow = list(row)

        active.append(listrow)
    wb.save(location)
    print("Success!!!! RS done")
    return results

def migration(date):
    cnx = connexion()
    # prepare a cursor object using cursor() method
    cursor = cnx.cursor()
    cursor.execute("select contractNo, ti 'type de client', TI_OLD 'ancien', TOTAL_AMOUNT as 'montant total', SWITCH_SERVICE_FEE as 'Migration Service', ITEM_AMMOUNT as 'TVA', "\
                  "FORMAT(op_time,'yyyy-MM-dd') as 'date opération', operator "\
                  "from ORDER_MASTER om "\
                  "inner join ORDER_CHARGE oc on om.ORDERSID=oc.ORDERSID "\
                  "where ORDER_TYPE=25 and  format(om.OP_TIME,'yyyy-MM-dd')= '"+date+"' and ITEM_NAME like '%TVA%' order by OPERATOR")

    # num_fields = len(cursor.description)
    field_names = [i[0] for i in cursor.description]
    # Mettre dans une liste
    results = cursor.fetchall()
    # Parcourir et print
    active = wb['Migration']
    active.append(field_names)
    for row in results:
        # make a row as List
        listrow = list(row)

        active.append(listrow)
    wb.save(location)
    print("Success!!!! Migration done")
    return results



def send_mail(fichier):
    """
       Sending mail with attachments from your Gmail account
       Args:
            nomFichier: file created with SQL Queries

    """
    fromaddr = "mbayedione10@gmail.com"
    toaddr = ["mbayedione10@gmail.com","nioulboy@gmail.com"]
    
    session = smtplib.SMTP('smtp.gmail.com', 587)
    print("creating SMTP session")

    # start TLS for security
    session.starttls()
    # Authentication
    session.login(fromaddr, "cmfoligafgmexlny")

    # instance of MIMEMultipart
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['Subject'] = "Vente du jour"
    body = """\
        Bonjour M. Dione
        Bonne reception du rapport journalier"""
    
    # open the file to be sent
    filename=nomFichier
    attachment = open(fichier,"rb")
    # instance of MIMEBase and named as p
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    print("changing the payload into encoded form")
    # encode into base64
    encoders.encode_base64(p)
    print("encode into base64")
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    for to_email in toaddr:

        msg['To'] = to_email
        print("storing sender, receiver address and body email")

        msg.attach(MIMEText(body, 'plain'))
        print("attach the body with the msg instance")
        
        msg.attach(p)
        print("add the instance 'p' to instance 'msg'")

        # Converts the Multipart msg into a string
        text = msg.as_string()
        
        session.sendmail(fromaddr, to_email, text)
        print("mail sent successfully")
    session.quit()

def addSummSheet(path):
    df = pd.ExcelFile(path)
    df1 = pd.read_excel(df, sheet_name= 'Abonnement')
    df2 = pd.read_excel(df, sheet_name= 'RO')
    df3 = pd.read_excel(df, sheet_name= 'RS')
    df4 = pd.read_excel(df, sheet_name= 'Migration')

    print("Saving sum of 'abonnement', 'RO', 'RS' and 'Migration' in Sheet")
    operator = df1.groupby(['operator'])
    op_montant = operator['montant total'].sum()
    operatorRO = df2.groupby(['operator'])
    op_montantRO = operatorRO['montant total'].sum()
    operatorRS = df3.groupby(['OPERATOR'])
    op_montantRS = operatorRS['montant total'].sum()
    operatorMig = df4.groupby(['operator'])
    op_montantMig = operatorMig['montant total'].sum()
    
    """
    Write results on the excel file
    """
    myfile = open(path, "rb+")
    # Create a Pandas Excel writer using openpyxl as the engine.
    writer = pd.ExcelWriter(myfile, engine='openpyxl')
    # try to open an existing workbook
    writer.book = load_workbook(myfile)
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    op_montant.to_excel(writer,sheet_name='Sheet',startcol=0,startrow=1)
    op_montantRO.to_excel(writer,sheet_name='Sheet',startcol=4,startrow=1)
    op_montantRS.to_excel(writer,sheet_name='Sheet',startcol=7,startrow=1)
    op_montantMig.to_excel(writer,sheet_name='Sheet',startcol=10,startrow=1)
    
    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
    writer.close()



def remove_old_files(backup_dir):
    """
    Removes the old archives in backup directory which exceed retention days.
    Args:
        backup_dir: Path to the backup directory.
    """

    date = dt.datetime.today()
    old_date = date + dt.timedelta(days=-10)
    for f in os.listdir(backup_dir):
         if dt.datetime.fromtimestamp(os.path.getmtime(os.path.join(backup_dir,f))) < old_date:
             os.remove(os.path.join(backup_dir, f))
             print("Removing old backup {0} after 10 retention days".format(f))
         else:
             print("Old backup {0} is not older then 10 days. Skipping removal...".format(f))


#TODO between creating excel file and send mails --> time.sleep(10)

def main():
    jour=today.strftime("%Y-%m-%d")
    abonnement(jour)
    rechargeObligatoire(jour)
    rechargeSupplementaire(jour)
    migration(jour)
    
    time.sleep(10)

    # Choose last file uploaded in Ventes Folder and Send email
    list_of_files = glob.iglob("E:\\SQLPython\\Ventes\\*.xlsx")  # * means all if need specific format then *.csv
    latest_file = sorted(list_of_files, key=os.path.getmtime, reverse=True)[:1]
    
    addSummSheet(latest_file[0])
    
    time.sleep(5)

    send_mail(latest_file[0])
    BACKUP_DIR = r"E:\SQLPython\Ventes"
    remove_old_files(backup_dir=BACKUP_DIR)

main()
