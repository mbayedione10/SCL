#NioulBoy
import glob
import os

import pyodbc
import datetime as dt
from openpyxl import Workbook, load_workbook

import  smtplib
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

#Create workbook
wb = Workbook()
# Add sheets to workbook
# Designate sheet name and position
sheet1 = wb.create_sheet('Anticipation',0)
sheet2 = wb.create_sheet('RO',1)
sheet3 = wb.create_sheet('RS',2)
sheet4 = wb.create_sheet('Hexpay',3)
sheet5 = wb.create_sheet('Transfert',4)


#method to return the previous month like YYYY-MM
def yearMonth():
  today = dt.date.today()
  first = today.replace(day=1)
  lastMonth = first - dt.timedelta(days=1)
  return lastMonth.strftime("%Y-%m")



date = yearMonth()

filename = r"E:\SQLPython\{0}-{1}{2}".format("RequeteSQL", date, ".xlsx")


# connect to MySQL Server SRV-DB
def connexion():
  conn = pyodbc.connect(Driver = '{SQL Server Native Client 11.0}',
                        Server = 'SRV-DB',
                        Database = 'Server Name',
                        user='login',
                        password="password"
                        )
  return conn




#--TRANSFERT
def listeTransfert(year):
  cnx = connexion()
  # prepare a cursor object using cursor() method
  cursor  = cnx.cursor()
  cursor.execute("select om.contractNo, om.METERNO, cc.ADDR, cc.SERVICE_LEVEL, "\
                 "om.OPERATOR, format(om.OP_TIME,'yyyy-MM-dd') as 'DATE', om.ENERGY, om.ADDTIONAL_ENERGY "\
                 "from ORDER_MASTER om LEFT JOIN CON_CONTRACT cc on om.METERNO=cc.METER_NO "\
                 "where om.ORDER_TYPE='28'and format(om.OP_TIME,'yyyy-MM')= '"+year+"' "\
                 "order by om.OP_TIME desc")

  #Mettre dans une liste
  results = cursor.fetchall()
  active = wb['Transfert']
  #Parcourir et print
  for row in results:
    #Replace None value with 0
    if row[6] == None:
        row[6] = 0
    if row[7] == None:
        row[7] = 0
    #make a row as List
    listrow = list(row)
    active.append(listrow)
  wb.save(filename)
  print("Success!!!! Transfert done")
  return results


#-- ANTICIPATION
def listeAnticipation(year):
  cnx = connexion()
  # prepare a cursor object using cursor() method
  cursor  = cnx.cursor()
  cursor.execute("select om.contractNo as 'Num Contrat',om.TI AS 'Type Client',"\
                 "format(om.FIXED_START_MONTH,'yyyy-MM') as 'Debut', format(om.FIXED_END_MONTH, 'yyyy-MM') as 'Fin', format(om.OP_TIME,'yyyy-MM') as 'Date de recharge'"\
                 "from ORDER_MASTER om where om.ORDER_TYPE='01' and om.VENDINGMETHOD='01' "\
                 "and format(om.OP_TIME,'yyyy-MM')<'"+year+"' and format(om.FIXED_START_MONTH,'yyyy-MM')<='"+year+"'"\
                 "and format(om.FIXED_END_MONTH, 'yyyy-MM')>='"+year+"'  order by om.OP_TIME")


  results = cursor.fetchall()
  active = wb['Anticipation']
  for row in results:
    #make a row as List 
    listrow = list(row)
    active.append(listrow)
  wb.save(filename)
  print("Success!!!! Anticipation done")
  return results


#-- RECHARGE SUPPLEMENTAIRE
def listeRS(year):
  cnx = connexion()
  # prepare a cursor object using cursor() method
  cursor  = cnx.cursor()
  cursor.execute("select om.contractNo as 'Num Contrat',om.TI AS 'Type Client',om.ADDTIONAL_ENERGY as 'Quantité Energie' "\
                 "from ORDER_MASTER om "\
                 "where om.ORDER_TYPE='01' and om.VENDINGMETHOD='02' and format(om.OP_TIME,'yyyy-MM')='"+year+"' ORDER BY om.OP_TIME")

 
  results = cursor.fetchall()
  active = wb['RS']
  for row in results:
    #make a row as List 
    listrow = list(row)
    active.append(listrow)
  wb.save(filename)
  print("Success!!!! RS done")
  return results


#-- RECHARGE OBLIGATOIRE
def listeRO(year):
  cnx = connexion()
  # prepare a cursor object using cursor() method
  cursor  = cnx.cursor()
  cursor.execute("select om.contractNo as 'Contract Num',om.TI AS 'Acc Type',format(om.FIXED_START_MONTH,'yyyy-MM') as 'Mois de Recharge'"\
                 "from ORDER_MASTER om "\
                 "where om.ORDER_TYPE='01' and om.VENDINGMETHOD='01' and format(om.OP_TIME,'yyyy-MM')='"+year+"' "\
                 "and format(om.FIXED_START_MONTH,'yyyy-MM') = '"+year+"' order by om.OP_TIME")


  results = cursor.fetchall()
  active = wb['RO']
  for row in results:
    #make a row as List 
    listrow = list(row)
    active.append(listrow)
  wb.save(filename)
  print("Success!!!! RO done")
  return results



#-- VENTE HEXPAY
def listeVenteHexpay(year):
  cnx = connexion()
  # prepare a cursor object using cursor() method
  cursor1  = cnx.cursor()
  cursor1.execute("select CAST(om.OP_TIME as date) as 'Jour', SUM(om.TOTAL_AMOUNT) as 'RS et RO'"\
                 "from ORDER_MASTER om "\
                 "where om.ORDER_TYPE='01' and om.VENDINGMETHOD='01'  and format(om.OP_TIME,'yyyy-MM')='"+year+"' AND "\
                 "om.POSID = '1658' group by CAST(om.OP_TIME as date) order by Jour")
  
  results = cursor1.fetchall()
  active = wb['Hexpay']
 
  for row in results:
    #make a row as List
    active['A1'] = "Recharge Obligatoire Hexpay"
    listrow = list(row)
    active.append(listrow)
    
  cursor2  = cnx.cursor()
  cursor2.execute("select CAST(om.OP_TIME as date) as 'Jour', SUM(om.TOTAL_AMOUNT) as 'RS et RO'"\
                 "from ORDER_MASTER om "\
                 "where om.ORDER_TYPE='01' and om.VENDINGMETHOD='02'  and format(om.OP_TIME,'yyyy-MM')='"+year+"' AND "\
                 "om.POSID = '1658' group by CAST(om.OP_TIME as date) order by Jour")

  
  results2 = cursor2.fetchall()
  active = wb['Hexpay']
 
  for row2 in results2:
    #make a row as List
    active['C33'] = "Recharge Supplémentaire Hexpay"
    listrow2 = [None]*2 + list(row2)
    active.append(listrow2)

  wb.save(filename)
  print("Success!!!! Vente Hexpay done")
  return results

#-- RESTANT PRISE EN CHARGE
def listePriseEnCharge():
  cnx = connexion()
  # prepare a cursor object using cursor() method
  cursor  = cnx.cursor()
  cursor.execute("select CON_NO,ACC_NO, METER_NO,ADDR, dy.HM,SERVICE_LEVEL, tel.SJHM as 'Phone', BALANCE  from CON_CONTRACT cc "
                 "left join da_yh dy on cc.ACC_NO=dy.HH Left join DA_YHLXR tel on cc.ACC_NO=tel.HH "\
                 "where BALANCE >0 and cc.CON_SIGN in ('01','03','04') order by BALANCE desc") 

  #Mettre dans une liste
  results = cursor.fetchall()
  active = wb['Prise en charge']
  #Parcourir et print
  for row in results:
    #make a row as List 
    listrow = list(row)
    active.append(listrow)
  wb.save(filename)
  print("Success!!!! Prise en charge done")
  return results



def sendMail():
    conn =smtplib.SMTP('smtp-mail.outlook.com',587)
    type(conn)
    conn.ehlo()
    conn.starttls()
    conn.login('SenderMail','password')
    receiver = "ReceiverMail"
    message = """
    Subject: Requete SQL Mensuel

    SQL Request sent successfully"""

    try:
        conn.sendmail('SenderMail',receiver,message,)
        print("Mail sent")
        conn.quit()
    except:
        print("An exception occurred")
        


"""
The Google APIs client library uses client_secrets.json files for storing the client_id, client_secret, and other OAuth 2.0 parameters.

A client_secrets.json file is a JSON formatted file containing the client ID, client secret, and other OAuth 2.0
save this file on directory
"""
#--Save file on drive
def fileToDrive():
    gauth = GoogleAuth()
    # Try to load saved client credentials
    gauth.LoadCredentialsFile("mycreds.txt")
    if gauth.credentials is None:
        # Authenticate if they're not there
        gauth.LocalWebserverAuth()
    elif gauth.access_token_expired:
        # Refresh them if expired
        gauth.Refresh()
    else:
        # Initialize the saved creds
        gauth.Authorize()
    # Save the current credentials to a file
    gauth.SaveCredentialsFile("mycreds.txt")
    drive = GoogleDrive(gauth)
    
    
    list_of_files = glob.iglob("E:\\SQLPython\\*.xlsx")  # * all specific format then *.xlsx
    latest_file = sorted(list_of_files, key=os.path.getmtime, reverse=True)[:1]
    
    for file in latest_file:
        print(file)
        file_metadata = {'title': os.path.basename(file)}
        file_drive = drive.CreateFile(metadata=file_metadata)
        file_drive.SetContentFile(file)
        file_drive.Upload()
    
        print("The file: " + file + " has been uploaded")
    
    print("All files have been uploaded")
    print("******************************************************************************")
    sendMail()
    



def main():
  anMois = yearMonth()
  
  print("Liste des recharges par anticipation")
  listeAnticipation(anMois)
  print("******************************************************************************")
  
  print("liste des recharges obligatoires")
  listeRO(anMois)
  print("******************************************************************************")
  
  print("liste des recharges supplementaires")
  listeRS(anMois)
  print("******************************************************************************")
  
  print("liste des ventes Hexpay")
  listeVenteHexpay(anMois)
  print("******************************************************************************")
  
  print("liste des transfert par operateurs")
  listeTransfert(anMois)
  print("******************************************************************************")

  print("liste des prises en charges restants")
  listePriseEnCharge()
  print("******************************************************************************")


  
  fileToDrive()



    
if __name__ == '__main__':
    main()
